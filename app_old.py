from flask import Flask, request, Response, render_template
from openpyxl import load_workbook
import io
import re
import logging
from datetime import datetime
import unicodedata
import tempfile

app = Flask(__name__)

# Configure logging
logging.basicConfig(
    filename='app.log',
    level=logging.DEBUG,
    format='%(asctime)s %(levelname)s: %(message)s'
)

# ---------- Config ----------
BIB_KEYS = [
    '020$a', '020$c', '040$a', '040$b', '040$c', '041$a', '082$a', '082$b', '100$a', '245$a', '245$b', '250$a',
    '260$a', '260$b', '260$c', '300$a', '300$b', '362$a', '942$c', '650$a', '700$a'
]
MULTI_SPLIT = '|'

# ---------- Helpers ----------
def norm(v: str) -> str:
    v = v.strip(" \t\n\r\0\x0B").replace('\xa0', ' ')
    v = unicodedata.normalize('NFKC', v)
    return ' '.join(v.lower().split())

def format_date(val) -> str:
    if isinstance(val, datetime):
        return val.strftime('%Y-%m-%d')
    val_str = str(val).strip(" \t\n\r\0\x0B").replace('\xa0', ' ')
    # Handle string dates (e.g., '2025-08-14 00:00:00' or '2025-08-14')
    if re.match(r'\d{4}-\d{2}-\d{2}( \d{2}:\d{2}:\d{2})?', val_str):
        try:
            dt = datetime.strptime(val_str, '%Y-%m-%d %H:%M:%S') if ' ' in val_str else datetime.strptime(val_str, '%Y-%m-%d')
            return dt.strftime('%Y-%m-%d')
        except ValueError:
            logging.error(f"Invalid date format: {val_str}")
            return val_str
    return val_str

def line_mrk_pairs(tag: str, pairs: list, ind1: str = '\\', ind2: str = '\\') -> str:
    parts = []
    for p in pairs:
        if not isinstance(p, list) or len(p) < 2:
            logging.error(f"Invalid pair for tag {tag}: {p}")
            continue
        code, val = p
        val_formatted = format_date(val)
        if not val_formatted:
            logging.error(f"Empty or null value for tag {tag}, code {code}: raw='{val}'")
            continue
        parts.append(f"${code}{val_formatted}")
        logging.debug(f"Added subfield for tag {tag}, code {code}: '{val_formatted}' (raw type: {type(val).__name__})")
    if not parts:
        logging.error(f"No valid parts for tag {tag}: {pairs}")
        return ''
    # Enforce subfield order for 952
    if tag == '952':
        subfield_order = ['p', 'd', 'o', 'e', 'g', 'A', 'B', 'c', 'C', 'x', 'y', 'a', 'b']
        sorted_parts = []
        for code in subfield_order:
            for part in parts:
                if part.startswith(f"${code}"):
                    sorted_parts.append(part)
                    break
        parts = sorted_parts
    result = f"={tag}  {ind1}{ind2}{''.join(parts)}"
    logging.debug(f"Generated MRK line for tag {tag}: {result}")
    return result

def line_mrk(tag: str, subs: dict, ind1: str = '\\', ind2: str = '\\') -> str:
    parts = []
    for code, val in subs.items():
        val_formatted = format_date(val)
        if not val_formatted:
            continue
        parts.append(f"${code}{val_formatted}")
        logging.debug(f"Added subfield for tag {tag}, code {code}: '{val_formatted}' (raw type: {type(val).__name__})")
    if not parts:
        return ''
    return f"={tag}  {ind1}{ind2}{''.join(parts)}"

def first_nonempty(c: str | None, v: str | None) -> str | None:
    v = format_date(v) if v is not None else None
    if c is not None and c != '':
        return c
    if v is not None and v != '':
        return v
    return c

def build_ldr() -> str:
    return "=LDR  00000nam a2200000Ia 4500"

def build_008(lang: str) -> str:
    date = datetime.now().strftime("%y%m%d")
    return f"=008  {date}s9999||||xx\\||||||||||||||\\||{lang}||"

def strip_all_spaces(s: str) -> str:
    return re.sub(r'\s+', '', s)

def is_row_empty(row: list) -> bool:
    for cell in row:
        val = format_date(cell) if cell is not None else ''
        if val:
            return False
    return True

@app.route('/', methods=['GET'])
def index():
    return render_template('index.html')

@app.route('/upload', methods=['POST'])
def upload_file():
    lang = request.form.get('lang', 'und').strip(" \t\n\r\0\x0B").replace('\xa0', ' ')

    # ---------- File Handling ----------
    if 'file' not in request.files or request.files['file'].filename == '':
        logging.error("No file uploaded")
        return render_template('index.html', error="Upload failed: No file selected."), 400

    file = request.files['file']
    if not file.filename.lower().endswith('.xlsx'):
        logging.error(f"Invalid file extension: {file.filename}")
        return render_template('index.html', error="Only .xlsx files are allowed."), 400

    try:
        workbook = load_workbook(file.stream, read_only=True)
        sheet = workbook.active
        rows = list(sheet.rows)
    except Exception as e:
        logging.error(f"File loading error: {str(e)}")
        return render_template('index.html', error=f"Error loading file: {str(e)}"), 400

    if not rows or len(rows) < 1:
        logging.error("No data rows in Excel file")
        return render_template('index.html', error="No data rows in Excel file."), 400

    # ---------- Read Headers ----------
    headers = {}
    raw_headers = {}
    first_row = [cell.value for cell in rows[0]]
    for col_idx, h in enumerate(first_row, 1):
        h = str(h) if h is not None else ''
        raw_headers[col_idx] = h
        clean = h.strip(" \t\n\r\0\x0B").replace('\xa0', ' ')
        if clean:
            headers[col_idx] = clean
    logging.debug(f"Raw Headers: {raw_headers}")
    rows = rows[1:]  # Skip header row

    # ---------- Identify 952 Columns ----------
    holding_cols = []
    for col_idx, raw in raw_headers.items():
        if not raw:
            continue
        # Case-sensitive match for 952 subfields
        if re.match(r'^952\$[A-Za-z0-9]$', raw):
            match = re.search(r'\$([A-Za-z0-9])', raw)
            if match:
                holding_cols.append({'col': col_idx, 'code': match.group(1)})
                logging.debug(f"Detected 952 subfield: column {col_idx}, code {match.group(1)}, raw header '{raw}'")
    logging.debug(f"Holding Columns: {holding_cols}")

    # ---------- Grouping ----------
    groups = {}
    recid = 1
    row_count = len(rows)
    last_row_index = row_count - 1

    for row_idx, row_cells in enumerate(rows):
        row = [cell.value for cell in row_cells]
        is_last_row = (row_idx == last_row_index)

        # Skip empty rows
        if is_row_empty(row):
            logging.debug(f"Row {row_idx}" + (" (Last)" if is_last_row else "") + " Skipped: Empty row")
            continue

        row_data = {}
        for col_idx, h in headers.items():
            cell = row[col_idx - 1] if col_idx - 1 < len(row) else ''
            row_data[h] = cell

        # Debug: Log raw and trimmed 952 data
        debug_952 = {}
        for hc in holding_cols:
            raw_val = row[hc['col'] - 1] if hc['col'] - 1 < len(row) else ''
            trimmed_val = format_date(raw_val)
            debug_952[hc['code']] = {'raw': raw_val, 'trimmed': trimmed_val, 'type': type(raw_val).__name__}
        logging.debug(f"Row {row_idx}" + (" (Last)" if is_last_row else "") + f" 952 Data: {debug_952}")

        # Fallback: Re-read last row directly from sheet if empty
        if is_last_row and not any(v['trimmed'] for v in debug_952.values()):
            logging.debug(f"Row {row_idx} (Last) appears empty for 952; re-reading directly")
            row_number = row_idx + 2  # Account for header row
            for hc in holding_cols:
                cell = sheet.cell(row=row_number, column=hc['col']).value
                trimmed_val = format_date(cell)
                debug_952[hc['code']] = {'raw': cell, 'trimmed': trimmed_val, 'type': type(cell).__name__}
            logging.debug(f"Row {row_idx} (Last) Re-read 952 Data: {debug_952}")

        # Grouping key
        key_parts = []
        for fn in BIB_KEYS:
            v = row_data.get(fn, '')
            v = format_date(v)
            if fn in ['650$a', '700$a']:
                vals = [x.strip(" \t\n\r\0\x0B").replace('\xa0', ' ') for x in v.split(MULTI_SPLIT)]
                if not any(x for x in vals):
                    norm_vals = ['__EMPTY__']
                else:
                    norm_vals = [norm(x) for x in vals if x]
                    norm_vals = norm_vals if norm_vals else ['__EMPTY__']
                for i, nv in enumerate(norm_vals, 1):
                    key_parts.append(f"{fn}_{i}:{nv}")
            else:
                key_parts.append(f"{fn}:{norm(v)}")
        key = '||'.join(key_parts) if key_parts else f"row-{recid}"
        recid += 1

        # Initialize group
        if key not in groups:
            groups[key] = {
                'biblio': {
                    '020': {'a': None, 'c': None},
                    '040': {'a': None, 'b': None, 'c': None},
                    '041': {'a': None},
                    '082': {'a': None, 'b': None},
                    '100': {'a': None},
                    '245': {'a': None, 'b': None},
                    '250': {'a': None},
                    '260': {'a': None, 'b': None, 'c': None},
                    '300': {'a': None, 'b': None},
                    '362': {'a': None},
                    '942': {'c': None},
                    '650': [],
                    '700': [],
                },
                'holdings_pairs': [],
            }

        # Repeatables
        if row_data.get('650$a'):
            groups[key]['biblio']['650'].extend(
                x for x in row_data['650$a'].split(MULTI_SPLIT)
                if x.strip(" \t\n\r\0\x0B").replace('\xa0', ' ')
            )
        if row_data.get('700$a'):
            groups[key]['biblio']['700'].extend(
                x for x in row_data['700$a'].split(MULTI_SPLIT)
                if x.strip(" \t\n\r\0\x0B").replace('\xa0', ' ')
            )

        # Non-repeatables
        groups[key]['biblio']['020']['a'] = first_nonempty(groups[key]['biblio']['020']['a'], row_data.get('020$a'))
        groups[key]['biblio']['020']['c'] = first_nonempty(groups[key]['biblio']['020']['c'], row_data.get('020$c'))
        groups[key]['biblio']['040']['a'] = first_nonempty(groups[key]['biblio']['040']['a'], row_data.get('040$a'))
        groups[key]['biblio']['040']['b'] = first_nonempty(groups[key]['biblio']['040']['b'], row_data.get('040$b'))
        groups[key]['biblio']['040']['c'] = first_nonempty(groups[key]['biblio']['040']['c'], row_data.get('040$c'))
        groups[key]['biblio']['041']['a'] = first_nonempty(groups[key]['biblio']['041']['a'], row_data.get('041$a'))
        groups[key]['biblio']['082']['a'] = first_nonempty(groups[key]['biblio']['082']['a'], row_data.get('082$a'))
        groups[key]['biblio']['082']['b'] = first_nonempty(groups[key]['biblio']['082']['b'], row_data.get('082$b'))
        groups[key]['biblio']['100']['a'] = first_nonempty(groups[key]['biblio']['100']['a'], row_data.get('100$a'))
        groups[key]['biblio']['245']['a'] = first_nonempty(groups[key]['biblio']['245']['a'], row_data.get('245$a'))
        groups[key]['biblio']['245']['b'] = first_nonempty(groups[key]['biblio']['245']['b'], row_data.get('245$b'))
        groups[key]['biblio']['250']['a'] = first_nonempty(groups[key]['biblio']['250']['a'], row_data.get('250$a'))
        groups[key]['biblio']['260']['a'] = first_nonempty(groups[key]['biblio']['260']['a'], row_data.get('260$a'))
        groups[key]['biblio']['260']['b'] = first_nonempty(groups[key]['biblio']['260']['b'], row_data.get('260$b'))
        groups[key]['biblio']['260']['c'] = first_nonempty(groups[key]['biblio']['260']['c'], row_data.get('260$c'))
        groups[key]['biblio']['300']['a'] = first_nonempty(groups[key]['biblio']['300']['a'], row_data.get('300$a'))
        groups[key]['biblio']['300']['b'] = first_nonempty(groups[key]['biblio']['300']['b'], row_data.get('300$b'))
        groups[key]['biblio']['362']['a'] = first_nonempty(groups[key]['biblio']['362']['a'], row_data.get('362$a'))
        groups[key]['biblio']['942']['c'] = first_nonempty(groups[key]['biblio']['942']['c'], row_data.get('942$c'))

        # Holdings
        pairs = []
        for hc in holding_cols:
            col = hc['col']
            code = hc['code']
            cell = row[col - 1] if col - 1 < len(row) else ''
            val = format_date(cell)
            if val:
                pairs.append([code, val])
                logging.debug(f"Row {row_idx}" + (" (Last)" if is_last_row else "") + f" Added 952${code}: '{val}' (raw type: {type(cell).__name__})")
        if is_last_row and not pairs:
            logging.debug(f"Row {row_idx} (Last) No Holdings from row; re-reading directly")
            row_number = row_idx + 2  # Account for header row
            pairs = []
            for hc in holding_cols:
                cell = sheet.cell(row=row_number, column=hc['col']).value
                val = format_date(cell)
                if val:
                    pairs.append([hc['code'], val])
                    logging.debug(f"Row {row_idx} (Last) Re-read Added 952${hc['code']}: '{val}' (raw type: {type(cell).__name__})")
        if pairs:
            groups[key]['holdings_pairs'].append(pairs)
            logging.debug(f"Row {row_idx}" + (" (Last)" if is_last_row else "") + f" Holdings Added: {pairs}")
        else:
            logging.debug(f"Row {row_idx}" + (" (Last)" if is_last_row else "") + f" No Holdings: {debug_952}")

    # Deduplicate repeatables
    for key in groups:
        groups[key]['biblio']['650'] = list(dict.fromkeys(groups[key]['biblio']['650']))
        groups[key]['biblio']['700'] = list(dict.fromkeys(groups[key]['biblio']['700']))

    # ---------- Build MRK ----------
    def generate_mrk():
        next_001 = 1
        for key, g in groups.items():
            id_ = f"{next_001:09d}"
            next_001 += 1

            rec_lang = g['biblio']['041']['a'].strip(" \t\n\r\0\x0B").replace('\xa0', ' ') if g['biblio']['041']['a'] else lang

            # Start a new record
            record = []
            record.append(build_ldr())
            record.append(f"=001  {id_}")
            record.append(build_008(rec_lang))

            fields = [
                ['020', {'a': g['biblio']['020']['a'], 'c': g['biblio']['020']['c']}],
                ['040', {'a': g['biblio']['040']['a'], 'b': g['biblio']['040']['b'], 'c': g['biblio']['040']['c']}],
                ['041', {'a': g['biblio']['041']['a']}],
                ['082', {'a': g['biblio']['082']['a'], 'b': g['biblio']['082']['b']}],
                ['100', {'a': g['biblio']['100']['a']}],
                ['245', {'a': g['biblio']['245']['a'], 'b': g['biblio']['245']['b']}],
                ['250', {'a': g['biblio']['250']['a']}],
                ['260', {'a': g['biblio']['260']['a'], 'b': g['biblio']['260']['b'], 'c': g['biblio']['260']['c']}],
                ['300', {'a': g['biblio']['300']['a'], 'b': g['biblio']['300']['b']}],
                ['362', {'a': g['biblio']['362']['a']}],
                ['942', {'c': g['biblio']['942']['c']}],
            ]

            for tag, subs in fields:
                subs = {k: v for k, v in subs.items() if v is not None and str(v).strip(" \t\n\r\0\x0B").replace('\xa0', ' ')}
                if subs:
                    l = line_mrk(tag, subs)
                    if l:
                        record.append(l)

            for val in g['biblio']['650']:
                l = line_mrk('650', {'a': val})
                if l:
                    record.append(l)
            for val in g['biblio']['700']:
                l = line_mrk('700', {'a': val})
                if l:
                    record.append(l)

            for pairs in g['holdings_pairs']:
                l = line_mrk_pairs('952', pairs)
                if l:
                    record.append(l)
                    logging.debug(f"Added 952 to MRK (ID {id_}): {l}")

            # Log the complete record
            logging.debug(f"Record (ID {id_}): {' | '.join(record)}")

            # Explicitly log last record's full content
            if id_ == f"{len(groups):09d}":
                logging.debug(f"Last Record (ID {id_}) Full Content: {'\n'.join(record)}")

            for line in record:
                yield line + '\n'
            yield '\n'

    # Write MRK to temporary file for debugging
    try:
        with tempfile.NamedTemporaryFile(delete=False, suffix='.mrk', mode='w', encoding='utf-8') as temp_file:
            for line in generate_mrk():
                temp_file.write(line)
            temp_file_path = temp_file.name
        logging.debug(f"MRK content written to temporary file: {temp_file_path}")
    except Exception as e:
        logging.error(f"Error writing to temporary file: {str(e)}")

    # ---------- Download ----------
    try:
        fname = f"export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.mrk"
        mrk_content = ''.join(generate_mrk())
        logging.debug(f"MRK Content Length: {len(mrk_content)} bytes")
        return Response(
            mrk_content,
            mimetype='text/plain; charset=UTF-8',
            headers={'Content-Disposition': f'attachment; filename="{fname}"'}
        )
    except Exception as e:
        logging.error(f"Output error: {str(e)}")
        return render_template('index.html', error="Error generating MRK file."), 500

if __name__ == '__main__':
    app.run(debug=True, host='0.0.0.0', port=5000)
