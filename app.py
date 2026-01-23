from flask import Flask, request, Response, render_template
from openpyxl import load_workbook
import io
import re
import logging
from datetime import datetime
import unicodedata
import tempfile
import os

app = Flask(__name__)

# Configure logging
logging.basicConfig(
    filename='app.log',
    level=logging.DEBUG,
    format='%(asctime)s %(levelname)s: %(message)s'
)

# ---------- Config ----------
BIB_KEYS = [
    '020$a', '020$c', '040$a', '040$b', '040$c', '040$d', '041$a', '082$a', '082$b', '100$a', '110$a', '245$a', '245$b', '246$a',
    '250$a', '260$a', '260$b', '260$c', '300$a', '300$b', '300$c', '300$e', '362$a', '365$a', '365$b', '365$c', '365$d', '365$e',
    '365$j', '490$a', '490$v', '500$a', '520$a', '521$a', '942$c', '856$u', '650$a', '650$x', '700$a', '710$a'
]
MULTI_SPLIT = '|'

# ---------- Helpers ----------
def norm(v: str) -> str:
    if v is None:
        return ''
    v = v.strip(" \t\n\r\0\x0B").replace('\xa0', ' ')
    v = unicodedata.normalize('NFKC', v)
    return ' '.join(v.lower().split())

def format_date(val) -> str:
    if val is None:
        return ''

    if isinstance(val, datetime):
        return val.strftime('%Y-%m-%d')
    
    val_str = str(val).strip(" \t\n\r\0\x0B").replace('\xa0', ' ')
    
    if not val_str or val_str.lower() == 'none':
        return ''

    if re.match(r'\d{4}-\d{2}-\d{2}( \d{2}:\d{2}:\d{2})?', val_str):
        try:
            dt = datetime.strptime(val_str, '%Y-%m-%d %H:%M:%S') if ' ' in val_str else datetime.strptime(val_str, '%Y-%m-%d')
            return dt.strftime('%Y-%m-%d')
        except ValueError:
            logging.error(f"Invalid date format: {val_str}")
            return val_str
            
    return val_str

def line_mrk_pairs(tag: str, pairs: list, ind1: str = '\\', ind2: str = '\\') -> str:
    parts = []
    for p in pairs:
        if not isinstance(p, list) or len(p) < 2:
            logging.error(f"Invalid pair for tag {tag}: {p}")
            continue
        code, val = p
        val_formatted = format_date(val)
        
        if not val_formatted:
            continue
            
        parts.append(f"${code}{val_formatted}")
        
    if not parts:
        return ''
        
    if tag == '952':
        subfield_order = ['p', 'd', 'o', 'e', 'g', '0', '1', '2', '4', '5', '7', 'f', 't', 'A', 'B', 'c', 'C', 't', '2', '8', 'w', 'x', 'z', 'y', 'a', 'b']
        sorted_parts = []
        for code in subfield_order:
            for part in parts:
                if part.startswith(f"${code}"):
                    sorted_parts.append(part)
        used_parts = set(sorted_parts)
        for part in parts:
            if part not in used_parts:
                sorted_parts.append(part)
        parts = sorted_parts
        
    result = f"={tag}  {ind1}{ind2}{''.join(parts)}"
    return result

def line_mrk(tag: str, subs: dict, ind1: str = '\\', ind2: str = '\\') -> str:
    parts = []
    for code, val in subs.items():
        val_formatted = format_date(val)
        if not val_formatted:
            continue
        parts.append(f"${code}{val_formatted}")
    if not parts:
        return ''
    return f"={tag}  {ind1}{ind2}{''.join(parts)}"

def first_nonempty(c: str | None, v: str | None) -> str | None:
    v = format_date(v) if v is not None else None
    if c is not None and c != '':
        return c
    if v is not None and v != '':
        return v
    return c

def build_ldr() -> str:
    return "=LDR  00000nam a2200000Ia 4500"

def build_008(lang: str) -> str:
    date = datetime.now().strftime("%y%m%d")
    return f"=008  {date}s9999||||xx\\||||||||||||||\\||{lang}||"

def is_row_empty(row: list) -> bool:
    for cell in row:
        val = format_date(cell) if cell is not None else ''
        if val:
            return False
    return True

@app.route('/', methods=['GET'])
def index():
    return render_template('index.html')

@app.route('/upload', methods=['POST'])
def upload_file():
    lang = request.form.get('lang', 'und').strip(" \t\n\r\0\x0B").replace('\xa0', ' ')

    # ---------- File Handling ----------
    if 'file' not in request.files or request.files['file'].filename == '':
        return render_template('index.html', error="Upload failed: No file selected."), 400

    file = request.files['file']
    if not file.filename.lower().endswith('.xlsx'):
        return render_template('index.html', error="Only .xlsx files are allowed."), 400

    try:
        workbook = load_workbook(file.stream, read_only=True)
        sheet = workbook.active
        rows = list(sheet.rows)
    except Exception as e:
        return render_template('index.html', error=f"Error loading file: {str(e)}"), 400

    if not rows or len(rows) < 1:
        return render_template('index.html', error="No data rows in Excel file."), 400

    # ---------- Read Headers ----------
    headers = {}
    raw_headers = {}
    first_row = [cell.value for cell in rows[0]]
    for col_idx, h in enumerate(first_row, 1):
        h = str(h) if h is not None else ''
        raw_headers[col_idx] = h
        clean = h.strip(" \t\n\r\0\x0B").replace('\xa0', ' ')
        if clean:
            headers[col_idx] = clean
    rows = rows[1:]

    # ---------- Identify 952 Columns ----------
    holding_cols = []
    for col_idx, raw in raw_headers.items():
        if not raw: continue
        if re.match(r'^952\$[A-Za-z0-9]$', raw):
            match = re.search(r'\$([A-Za-z0-9])', raw)
            if match:
                holding_cols.append({'col': col_idx, 'code': match.group(1)})

    # ---------- Grouping ----------
    groups = {}
    recid = 1
    row_count = len(rows)
    last_row_index = row_count - 1

    for row_idx, row_cells in enumerate(rows):
        row = [cell.value for cell in row_cells]
        is_last_row = (row_idx == last_row_index)

        if is_row_empty(row):
            continue

        row_data = {}
        for col_idx, h in headers.items():
            cell = row[col_idx - 1] if col_idx - 1 < len(row) else ''
            row_data[h] = cell

        # Fallback: Re-read 952 if needed
        pairs = []
        for hc in holding_cols:
            col = hc['col']
            code = hc['code']
            cell = row[col - 1] if col - 1 < len(row) else ''
            val = format_date(cell)
            if val:
                pairs.append([code, val])
        
        if is_last_row and not pairs:
            row_number = row_idx + 2
            for hc in holding_cols:
                cell = sheet.cell(row=row_number, column=hc['col']).value
                val = format_date(cell)
                if val:
                    pairs.append([hc['code'], val])

        # Grouping key
        key_parts = []
        for fn in BIB_KEYS:
            v = row_data.get(fn, '')
            v = format_date(v)
            if fn in ['650$a', '700$a', '710$a']:
                vals = [x.strip(" \t\n\r\0\x0B").replace('\xa0', ' ') for x in v.split(MULTI_SPLIT)]
                norm_vals = [norm(x) for x in vals if x] if any(x for x in vals) else ['__EMPTY__']
                for i, nv in enumerate(norm_vals, 1):
                    key_parts.append(f"{fn}_{i}:{nv}")
            else:
                key_parts.append(f"{fn}:{norm(v)}")
        key = '||'.join(key_parts) if key_parts else f"row-{recid}"
        recid += 1

        # Initialize group
        if key not in groups:
            groups[key] = {
                'biblio': {
                    '020': [], 
                    '040': {'a': None, 'b': None, 'c': None, 'd': None},
                    '041': {'a': None},
                    '082': {'a': None, 'b': None},
                    '100': {'a': None},
                    '110': {'a': None},
                    '245': {'a': None, 'b': None},
                    '246': {'a': None},
                    '250': {'a': None},
                    '260': {'a': None, 'b': None, 'c': None},
                    '300': {'a': None, 'b': None, 'c': None, 'e': None},
                    '362': {'a': None},
                    '365': {'a': None, 'b': None, 'c': None, 'd': None, 'e': None, 'j': None},
                    '490': {'a': None, 'v': None},
                    '500': {'a': None},
                    '520': {'a': None},
                    '521': {'a': None},
                    '856': {'u': None},
                    '942': {'c': None},
                    '650': [],
                    '700': [],
                    '710': [],
                },
                'holdings_pairs': [],
            }

        # Handle 020 (ISBN) Split logic
        val_020a = row_data.get('020$a')
        val_020c = row_data.get('020$c') # Terms of availability
        if val_020a:
            # Split and clean
            isbns = [x.strip(" \t\n\r\0\x0B").replace('\xa0', ' ') for x in val_020a.split(MULTI_SPLIT)]
            for idx, isbn in enumerate(isbns):
                if isbn:
                    # FIX: Attach 'c' ONLY to the first ISBN (idx == 0)
                    current_c = val_020c if idx == 0 else None
                    groups[key]['biblio']['020'].append({'a': isbn, 'c': current_c})

        # --- UPDATED 650 LOGIC START ---
        val_650a = row_data.get('650$a')
        val_650x = row_data.get('650$x')

        if val_650a:
            # Split both fields
            list_a = [x.strip(" \t\n\r\0\x0B").replace('\xa0', ' ') for x in val_650a.split(MULTI_SPLIT)]
            list_x = [x.strip(" \t\n\r\0\x0B").replace('\xa0', ' ') for x in val_650x.split(MULTI_SPLIT)] if val_650x else []
            
            # Combine them by index
            for idx, sub_a in enumerate(list_a):
                if sub_a:
                    # Get corresponding x if it exists, else None
                    sub_x = list_x[idx] if idx < len(list_x) and list_x[idx] else None
                    groups[key]['biblio']['650'].append({'a': sub_a, 'x': sub_x})
        # --- UPDATED 650 LOGIC END ---

        # Repeatables
        if row_data.get('700$a'):
            groups[key]['biblio']['700'].extend(x for x in row_data['700$a'].split(MULTI_SPLIT) if x.strip())
        if row_data.get('710$a'):
            groups[key]['biblio']['710'].extend(x for x in row_data['710$a'].split(MULTI_SPLIT) if x.strip())

        # Non-repeatables
        groups[key]['biblio']['040']['a'] = first_nonempty(groups[key]['biblio']['040']['a'], row_data.get('040$a'))
        groups[key]['biblio']['040']['b'] = first_nonempty(groups[key]['biblio']['040']['b'], row_data.get('040$b'))
        groups[key]['biblio']['040']['c'] = first_nonempty(groups[key]['biblio']['040']['c'], row_data.get('040$c'))
        groups[key]['biblio']['040']['d'] = first_nonempty(groups[key]['biblio']['040']['d'], row_data.get('040$d'))
        groups[key]['biblio']['041']['a'] = first_nonempty(groups[key]['biblio']['041']['a'], row_data.get('041$a'))
        groups[key]['biblio']['082']['a'] = first_nonempty(groups[key]['biblio']['082']['a'], row_data.get('082$a'))
        groups[key]['biblio']['082']['b'] = first_nonempty(groups[key]['biblio']['082']['b'], row_data.get('082$b'))
        groups[key]['biblio']['100']['a'] = first_nonempty(groups[key]['biblio']['100']['a'], row_data.get('100$a'))
        groups[key]['biblio']['110']['a'] = first_nonempty(groups[key]['biblio']['110']['a'], row_data.get('110$a'))
        groups[key]['biblio']['245']['a'] = first_nonempty(groups[key]['biblio']['245']['a'], row_data.get('245$a'))
        groups[key]['biblio']['245']['b'] = first_nonempty(groups[key]['biblio']['245']['b'], row_data.get('245$b'))
        groups[key]['biblio']['246']['a'] = first_nonempty(groups[key]['biblio']['246']['a'], row_data.get('246$a'))
        groups[key]['biblio']['250']['a'] = first_nonempty(groups[key]['biblio']['250']['a'], row_data.get('250$a'))
        groups[key]['biblio']['260']['a'] = first_nonempty(groups[key]['biblio']['260']['a'], row_data.get('260$a'))
        groups[key]['biblio']['260']['b'] = first_nonempty(groups[key]['biblio']['260']['b'], row_data.get('260$b'))
        groups[key]['biblio']['260']['c'] = first_nonempty(groups[key]['biblio']['260']['c'], row_data.get('260$c'))
        groups[key]['biblio']['300']['a'] = first_nonempty(groups[key]['biblio']['300']['a'], row_data.get('300$a'))
        groups[key]['biblio']['300']['b'] = first_nonempty(groups[key]['biblio']['300']['b'], row_data.get('300$b'))
        groups[key]['biblio']['300']['c'] = first_nonempty(groups[key]['biblio']['300']['c'], row_data.get('300$c'))
        groups[key]['biblio']['300']['e'] = first_nonempty(groups[key]['biblio']['300']['e'], row_data.get('300$e'))
        groups[key]['biblio']['362']['a'] = first_nonempty(groups[key]['biblio']['362']['a'], row_data.get('362$a'))
        groups[key]['biblio']['365']['a'] = first_nonempty(groups[key]['biblio']['365']['a'], row_data.get('365$a'))
        groups[key]['biblio']['365']['b'] = first_nonempty(groups[key]['biblio']['365']['b'], row_data.get('365$b'))
        groups[key]['biblio']['365']['c'] = first_nonempty(groups[key]['biblio']['365']['c'], row_data.get('365$c'))
        groups[key]['biblio']['365']['d'] = first_nonempty(groups[key]['biblio']['365']['d'], row_data.get('365$d'))
        groups[key]['biblio']['365']['e'] = first_nonempty(groups[key]['biblio']['365']['e'], row_data.get('365$e'))
        groups[key]['biblio']['365']['j'] = first_nonempty(groups[key]['biblio']['365']['j'], row_data.get('365$j'))
        groups[key]['biblio']['490']['a'] = first_nonempty(groups[key]['biblio']['490']['a'], row_data.get('490$a'))
        groups[key]['biblio']['490']['v'] = first_nonempty(groups[key]['biblio']['490']['v'], row_data.get('490$v'))
        groups[key]['biblio']['500']['a'] = first_nonempty(groups[key]['biblio']['500']['a'], row_data.get('500$a'))
        groups[key]['biblio']['520']['a'] = first_nonempty(groups[key]['biblio']['520']['a'], row_data.get('520$a'))
        groups[key]['biblio']['521']['a'] = first_nonempty(groups[key]['biblio']['521']['a'], row_data.get('521$a'))
        groups[key]['biblio']['856']['u'] = first_nonempty(groups[key]['biblio']['856']['u'], row_data.get('856$u'))
        groups[key]['biblio']['942']['c'] = first_nonempty(groups[key]['biblio']['942']['c'], row_data.get('942$c'))

        if pairs:
            groups[key]['holdings_pairs'].append(pairs)

    # Deduplicate repeatables
    for key in groups:
        groups[key]['biblio']['700'] = list(dict.fromkeys(groups[key]['biblio']['700']))
        groups[key]['biblio']['710'] = list(dict.fromkeys(groups[key]['biblio']['710']))
        
        # Deduplicate 020
        unique_020 = []
        seen_isbns = set()
        for item in groups[key]['biblio']['020']:
            # We filter by ISBN ('a') only. 
            # If the first occurrence has 'c', it is kept. 
            # If a later duplicate has 'c', it is ignored (which is usually desired behavior).
            if item['a'] not in seen_isbns:
                unique_020.append(item)
                seen_isbns.add(item['a'])
        groups[key]['biblio']['020'] = unique_020

        # --- UPDATED 650 DEDUPLICATION ---
        unique_650 = []
        seen_650 = set()
        for item in groups[key]['biblio']['650']:
            # Create a tuple of (a, x) for unique checking
            pair_key = (item['a'], item['x'])
            if pair_key not in seen_650:
                unique_650.append(item)
                seen_650.add(pair_key)
        groups[key]['biblio']['650'] = unique_650
        # ---------------------------------

    # ---------- Build MRK ----------
    def generate_mrk():
        next_001 = 1
        for key, g in groups.items():
            id_ = f"{next_001:09d}"
            next_001 += 1

            rec_lang = g['biblio']['041']['a'].strip(" \t\n\r\0\x0B").replace('\xa0', ' ') if g['biblio']['041']['a'] else lang

            record = []
            record.append(build_ldr())
            record.append(f"=001  {id_}")
            record.append(build_008(rec_lang))

            fields = [
                ['040', {'a': g['biblio']['040']['a'], 'b': g['biblio']['040']['b'], 'c': g['biblio']['040']['c'], 'd': g['biblio']['040']['d']}],
                ['041', {'a': g['biblio']['041']['a']}],
                ['082', {'a': g['biblio']['082']['a'], 'b': g['biblio']['082']['b']}],
                ['100', {'a': g['biblio']['100']['a']}],
                ['110', {'a': g['biblio']['110']['a']}],
                ['245', {'a': g['biblio']['245']['a'], 'b': g['biblio']['245']['b']}],
                ['246', {'a': g['biblio']['246']['a']}],
                ['250', {'a': g['biblio']['250']['a']}],
                ['260', {'a': g['biblio']['260']['a'], 'b': g['biblio']['260']['b'], 'c': g['biblio']['260']['c']}],
                ['300', {'a': g['biblio']['300']['a'], 'b': g['biblio']['300']['b'], 'c': g['biblio']['300']['c'], 'e': g['biblio']['300']['e']}],
                ['362', {'a': g['biblio']['362']['a']}],
                ['365', {'a': g['biblio']['365']['a'], 'b': g['biblio']['365']['b'], 'c': g['biblio']['365']['c'], 'd': g['biblio']['365']['d'], 'e': g['biblio']['365']['e'], 'j': g['biblio']['365']['j']}],
                ['490', {'a': g['biblio']['490']['a'], 'v': g['biblio']['490']['v']}],
                ['500', {'a': g['biblio']['500']['a']}],
                ['520', {'a': g['biblio']['520']['a']}],
                ['521', {'a': g['biblio']['521']['a']}],
                ['856', {'u': g['biblio']['856']['u']}],
                ['942', {'c': g['biblio']['942']['c']}],
            ]

            # Generate 020 loop
            for item in g['biblio']['020']:
                subs = {'a': item['a']}
                if item['c']:
                    subs['c'] = item['c']
                l = line_mrk('020', subs)
                if l: record.append(l)

            for tag, subs in fields:
                cleaned_subs = {}
                for k, v in subs.items():
                    fmt = format_date(v)
                    if fmt:
                        cleaned_subs[k] = fmt
                
                if cleaned_subs:
                    l = line_mrk(tag, cleaned_subs)
                    if l:
                        record.append(l)

            for val in g['biblio']['650']:
                l = line_mrk('650', {'a': val})
                if l: record.append(l)
            for val in g['biblio']['700']:
                l = line_mrk('700', {'a': val})
                if l: record.append(l)
            for val in g['biblio']['710']:
                l = line_mrk('710', {'a': val})
                if l: record.append(l)

            for pairs in g['holdings_pairs']:
                l = line_mrk_pairs('952', pairs)
                if l: record.append(l)

            for line in record:
                yield line + '\n'
            yield '\n'

    # ---------- Download ----------
    try:
        fname = f"export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.mrk"
        mrk_content = ''.join(generate_mrk())
        return Response(
            mrk_content,
            mimetype='text/plain; charset=UTF-8',
            headers={'Content-Disposition': f'attachment; filename="{fname}"'}
        )
    except Exception as e:
        logging.error(f"Output error: {str(e)}")
        return render_template('index.html', error="Error generating MRK file."), 500

if __name__ == '__main__':
    port = int(os.environ.get("PORT", 10000))
    #app.run(host='0.0.0.0', port=port)
